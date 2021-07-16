import openpyxl
from pathlib import Path
from collections import defaultdict
from Node import Node
from copy import deepcopy


def clean(returnString):
	if any(element in returnString for element in ['D.F.', 'DF', 'CDMX', 'Zona']):
		returnString = 'México'

	returnString = returnString.replace("á", "a")
	returnString = returnString.replace("é", "e")
	returnString = returnString.replace("í", "i")
	returnString = returnString.replace("ó", "o")
	returnString = returnString.replace("ú", "u")

	returnString = returnString.rsplit(",")[0]
	# returnString = returnString.rsplit(" ")[0]

	return returnString


class ObtenerInformacion:
	def __init__(self):
		pass

	xlsx_file = Path('Rutas_Outbound.xlsx')
	wb_obj = openpyxl.load_workbook(xlsx_file)

	Ruta_ = wb_obj['Ruta']

	row = Ruta_.max_row
	# columnaFecha = 1
	COLUMNA_NUMERO_TRANSPORTE = 2
	COLUMNA_ORIGEN = 4
	COLUMNA_DESTINO = 5
	COLUMNA_DISTANCIA = 16
	COLUMNA_TONELADAS = 12
	COLUMNA_VIAJES = 13

	ciudadesCambiadas = {}
	sonMismaCiudad = {}

	rutas = {}
	rutasHelper = {}
	rutasNodos = {}

	arbol = []
	excel = []

	DISTANCIA_MINIMA_MISMA_CIUDAD = 60
	DISTANCIA_MINIMA_DESTINO_ORIGEN = 60
	DISTANCIA_MAXIMA_CIRCUITO = 4600  # la ruta con distancia maxima es 2292km
	DEPTH_MAXIMA_CIRCUITO = 5

	@staticmethod
	def Cosita(self):
		a = []
		for i in range(2, self.row + 1):
			if (self.Ruta_.cell(i, self.COLUMNA_ORIGEN).value not in [0, '0', "", " ", None, "(blank)"] and
					self.Ruta_.cell(i, self.COLUMNA_DESTINO).value not in ['a recogido', "", " ", None, "(blank)"] and
					"Radial" not in self.Ruta_.cell(i, self.COLUMNA_DESTINO).value) and \
					(clean(self.Ruta_.cell(i, self.COLUMNA_DESTINO).value) not in a):
				a.append(clean(self.Ruta_.cell(i, self.COLUMNA_DESTINO).value))
		a.sort()
		print("[" + ", ".join(a) + "]")

	@staticmethod
	def CrearDiccionarios(self):
		for i in range(2, self.row + 1):
			if (self.Ruta_.cell(i, self.COLUMNA_ORIGEN).value not in [0, '0', "", " ", None, "(blank)"] and
					self.Ruta_.cell(i, self.COLUMNA_DESTINO).value not in ['a recogido', "", " ", None, "(blank)"] and
					"Radial" not in self.Ruta_.cell(i, self.COLUMNA_DESTINO).value):

				origen = clean(self.Ruta_.cell(i, self.COLUMNA_ORIGEN).value)
				destino = clean(self.Ruta_.cell(i, self.COLUMNA_DESTINO).value)

				ObtenerInformacion.CreacionMismaCiudad(self, i, origen, destino)

				ObtenerInformacion.CreacionRutas(self, i, origen, destino)

				self.excel.append([origen, destino, self.Ruta_.cell(i, self.COLUMNA_NUMERO_TRANSPORTE).value,
								   float(self.Ruta_.cell(i, self.COLUMNA_TONELADAS).value),
								   float(self.Ruta_.cell(i, self.COLUMNA_DISTANCIA).value),
								   float(self.Ruta_.cell(i, self.COLUMNA_VIAJES).value)])

		ObtenerInformacion.CrearDiccionarios2(self)
		ObtenerInformacion.ImprimirMismaCiudad(self)

	@staticmethod
	def CrearDiccionarios2(self):
		for rutaXL in self.excel:
			#origen = rutaXL[0]
			#destino = rutaXL[1]
			origen = rutaXL[1] # ahora se van a crear partes del diccionario con los destino como origenes
			numero = rutaXL[2]
			toneladas = rutaXL[3]
			distancia = rutaXL[4]
			viajes = rutaXL[5]

			for destino in self.sonMismaCiudad.get(origen, []):
				ruta = [numero, destino, toneladas, distancia]

				node = Node(destino, toneladas, distancia, origen, viajes)

				if origen not in self.rutas:
					self.rutas[origen] = [ruta]
					self.rutasHelper[origen] = [destino]
					self.rutasNodos[origen] = [node]
				elif destino not in self.rutasHelper.get(origen, [None, None]) and destino != origen:
					self.rutas[origen].append(ruta)
					self.rutasHelper[origen].append(destino)
					self.rutasNodos[origen].append(node)
				if destino in self.rutasHelper.get(origen, [None]):
					([x for x in self.rutasNodos[origen] if x.ciudadDestino == destino])[0].viajes += node.viajes
					([x for x in self.rutasNodos[origen] if x.ciudadDestino == destino])[0].toneladas += node.toneladas

	@staticmethod
	def ImprimirMismaCiudad(self):
		print("DESTINOS SUFICIENTEMENTE CERCA PARA SER CONSIDERADOS COMO EL MISMO LUGAR")
		for k, v in self.sonMismaCiudad.items():
			print(k + ": [" + ", ".join(v) + "]")
		print("\n")

	@staticmethod
	def CreacionRutas(self, i, origen, destino):
		ruta = [self.Ruta_.cell(i, self.COLUMNA_NUMERO_TRANSPORTE).value,
				destino,
				float(self.Ruta_.cell(i, self.COLUMNA_TONELADAS).value),
				float(self.Ruta_.cell(i, self.COLUMNA_DISTANCIA).value)]

		node = Node(destino,
					float(self.Ruta_.cell(i, self.COLUMNA_TONELADAS).value),
					float(self.Ruta_.cell(i, self.COLUMNA_DISTANCIA).value),
					origen,
					float(self.Ruta_.cell(i, self.COLUMNA_VIAJES).value))

		if origen not in self.rutas:
			self.rutas[origen] = [ruta]
			self.rutasHelper[origen] = [destino]
			self.rutasNodos[origen] = [node]
		elif destino not in self.rutasHelper.get(origen, [None, None]) and destino != origen:
			self.rutas[origen].append(ruta)
			self.rutasHelper[origen].append(destino)
			self.rutasNodos[origen].append(node)
		if destino in self.rutasHelper.get(origen, [None]):
			([x for x in self.rutasNodos[origen] if x.ciudadDestino == destino])[0].viajes += node.viajes
			([x for x in self.rutasNodos[origen] if x.ciudadDestino == destino])[0].toneladas += node.toneladas

	@staticmethod
	def CreacionMismaCiudad(self, i, origen, destino):
		if self.Ruta_.cell(i, self.COLUMNA_DISTANCIA).value < self.DISTANCIA_MINIMA_MISMA_CIUDAD:

			if origen not in self.sonMismaCiudad:
				self.sonMismaCiudad[origen] = [destino]
			elif destino not in self.sonMismaCiudad[origen]:
				self.sonMismaCiudad[origen].append(destino)

			if destino not in self.sonMismaCiudad:
				self.sonMismaCiudad[destino] = [origen]
			elif origen not in self.sonMismaCiudad[destino]:
				self.sonMismaCiudad[destino].append(origen)

	@staticmethod
	def CreacionMismaCiudad2(self):
		xlsx_file_2 = Path('Rutas_Cemix_Origen_TodosDestinos.xlsx')
		wb_obj_2 = openpyxl.load_workbook(xlsx_file_2)

		Ruta_2_ = wb_obj_2['Sheet1']

		row_2 = Ruta_2_.max_row

		COLUMNA_ORIGEN_2 = 1
		COLUMNA_DESTINO_2 = 2
		COLUMNA_DISTANCIA_2 = 3

		for i in range(2, row_2 + 1):
			origen = clean(Ruta_2_.cell(i, COLUMNA_ORIGEN_2).value)
			destino = clean(Ruta_2_.cell(i, COLUMNA_DESTINO_2).value)

			if Ruta_2_.cell(i, COLUMNA_DISTANCIA_2).value < self.DISTANCIA_MINIMA_MISMA_CIUDAD:
				if origen not in self.sonMismaCiudad:
					self.sonMismaCiudad[origen] = [destino]
				elif destino not in self.sonMismaCiudad[origen]:
					self.sonMismaCiudad[origen].append(destino)

				if destino not in self.sonMismaCiudad:
					self.sonMismaCiudad[destino] = [origen]
				elif origen not in self.sonMismaCiudad[destino]:
					self.sonMismaCiudad[destino].append(origen)

	@staticmethod
	def CrearArbolRutas(self, nodoPapa, esInicio=False):
		if esInicio:
			for k in self.rutas.keys():
				n = Node()
				n.ciudadFinal = k
				n.viajes = 999999999
				n.toneladas = 999999999
				self.arbol.append(n)
				self.CrearArbolRutas(self, self.arbol[-1])
			return

		if nodoPapa.ciudadDestino is not None and nodoPapa.ciudadDestino != "":
			try:
				lista = self.rutasNodos[nodoPapa.ciudadDestino]
			except:
				lista = self.rutasNodos.get(self.sonMismaCiudad.get(nodoPapa.ciudadDestino, [None])[0], [])
		else:
			lista = self.rutasNodos[nodoPapa.ciudadFinal]

		for ruta in lista:
			ruta_ = deepcopy(ruta)
			nodoPapa.hijos.append(ruta_)

			ruta_.nodoPadre = nodoPapa
			ruta_.ciudadFinal = nodoPapa.ciudadFinal

			ruta_.depth = nodoPapa.depth + 1
			ruta_.distanciaEfectiva = nodoPapa.distanciaEfectiva + ruta_.distanciaRuta

			if ruta_.ciudadDestino in ruta_.ciudadesPrevias or any(
					i in self.sonMismaCiudad.get(ruta_.ciudadDestino, [None]) for i in ruta_.ciudadesPrevias):
				del nodoPapa.hijos[-1]
				return

			ruta_.ciudadesPrevias = deepcopy(nodoPapa.ciudadesPrevias)
			ruta_.ciudadesPrevias.append(nodoPapa.ciudadDestino)

			if ruta_.ciudadDestino == ruta_.ciudadFinal or ruta_.ciudadFinal in self.sonMismaCiudad.get(
					ruta_.ciudadDestino, [None]):

				ruta_.nodoPadre.viajesMax = ruta_.nodoPadre.viajesMax + ruta_.viajes if ruta_.nodoPadre.viajesMax + ruta_.viajes < ruta_.nodoPadre.viajes else ruta_.nodoPadre.viajes
				ruta_.nodoPadre.toneladasMax = ruta_.nodoPadre.toneladasMax + ruta_.toneladas if ruta_.nodoPadre.toneladasMax + ruta_.toneladas < ruta_.nodoPadre.toneladas else ruta_.nodoPadre.toneladas

				if ruta_.ciudadFinal in self.sonMismaCiudad.get(ruta_.ciudadDestino, [None]):
					ruta.ciudadDestino += "(" + ruta_.ciudadFinal + ")"
				ruta_.viajesMax = ruta_.viajes
				ruta_.toneladasMax = ruta_.toneladas
				return

			if ruta_.depth > self.DEPTH_MAXIMA_CIRCUITO or ruta_.distanciaEfectiva > self.DISTANCIA_MAXIMA_CIRCUITO:
				del nodoPapa.hijos[-1]
				return

			ruta_.nodoPadre.viajesMax = ruta_.nodoPadre.viajesMax + ruta_.viajes if ruta_.nodoPadre.viajesMax + ruta_.viajes < ruta_.nodoPadre.viajes else ruta_.nodoPadre.viajes
			ruta_.nodoPadre.toneladasMax = ruta_.nodoPadre.toneladasMax + ruta_.toneladas if ruta_.nodoPadre.toneladasMax + ruta_.toneladas < ruta_.nodoPadre.toneladas else ruta_.nodoPadre.toneladas

			self.CrearArbolRutas(self, ruta_)

		if nodoPapa is not None and (len(nodoPapa.hijos) == 0 or nodoPapa.hijos.count(None) == len(nodoPapa.hijos)) \
				and nodoPapa.nodoPadre is not None:
			nodoPapa.nodoPadre.hijos.remove(nodoPapa)

	@staticmethod
	def LimpiarArbolRutas(self, nodo, esInicio=False):
		if esInicio:
			for origen in self.arbol:
				self.LimpiarArbolRutas(self, origen)
			return

		if nodo is None:
			del nodo
			return

		for hijo in nodo.hijos:
			self.LimpiarArbolRutas(self, hijo)

		if (len(nodo.hijos) == 0 or nodo.hijos.count(None) == len(nodo.hijos)) and nodo.nodoPadre is not None and not (
				nodo.ciudadDestino == nodo.ciudadFinal or nodo.ciudadFinal in self.sonMismaCiudad.get(
				nodo.ciudadDestino, [None])):
			nodo.nodoPadre.hijos = [x if x is not nodo else None for x in nodo.nodoPadre.hijos]
			return

		nodo.hijos = [x for x in nodo.hijos if x is not None]

	@staticmethod
	def Backpropagate(self, nodo, esInicio=False):
		if esInicio:
			for origen in self.arbol:
				self.Backpropagate(self, origen)
			return

		if len(nodo.hijos) != 0:
			nodo.viajesMax = 0
			nodo.toneladasMax = 0
			for hijo in nodo.hijos:
				ObtenerInformacion.Backpropagate(self, hijo)
				nodo.viajesMax += hijo.viajesMax
				nodo.toneladasMax += hijo.toneladasMax
			nodo.viajesMax = min(nodo.viajesMax, nodo.viajes)
			nodo.toneladasMax = min(nodo.toneladasMax, nodo.toneladas)
		else:
			nodo.viajesMax = nodo.viajes
			nodo.toneladasMax = nodo.toneladas


	@staticmethod
	def LimpiarArbolNone(self, nodo, esInicio=False):
		if esInicio:
			for origen in self.arbol:
				self.LimpiarArbolNone(self, origen)
			return

		if nodo is None:
			return

		nodo.hijos = [x for x in nodo.hijos if x is not None]

		for hijo in nodo.hijos:
			self.LimpiarArbolNone(self, hijo)

		nodo.hijos = [x for x in nodo.hijos if x is not None]

	@staticmethod
	def ImprimirRutasKilometro(self):
		for k in self.rutas.keys():
			print(k)
			printed = []
			for v in self.rutas[k]:
				if v[1] not in printed and self.sonMismaCiudad.get(v[1], [None])[0] not in printed and \
						self.sonMismaCiudad.get(v[1], [None])[0] != k:
					print("\t" + str(v[1]) + ": " + str(v[3]) + "km")
					printed.append(v[1])

	@staticmethod
	def ImprimirRutas(self):
		print(self.rutas)

	@staticmethod
	def ImprimirRutas(self, ciudad):
		print(self.rutas[ciudad])
